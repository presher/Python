from pdf2image import convert_from_path
import pytesseract
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

#add path and output locations
pdf_path = "Documents/python/pdf to excel/June.pdf"

# Generate filename with month and year
def generate_filename(base="Ministering_Assignments"):
    month_year = datetime.now().strftime("%B_%Y")  # e.g., May_2025
    return f"{base}_{month_year}.xlsx"

# Convert PDF pages to images and extract OCR text
text_lines = []
images = convert_from_path(pdf_path)
for img in images:
    ocr_text = pytesseract.image_to_string(img)
    text_lines.extend(ocr_text.splitlines())

# Track presidency changes and break blocks by presidency member appearance
companionship_blocks = []
current_header = ""
presidency_member = ""
last_presidency = ""
district_counter = 1
current_block = []

header_displayed = set()

for line in text_lines:
    line = line.strip()

    presidency_match = re.search(r'Presidency Member:\s+([A-Za-z,\s]+)', line)
    if presidency_match:
        # Save previous block if non-empty
        if current_block:
            companionship_blocks.append((current_header, list(current_block)))
            current_block = []

        presidency_member = presidency_match.group(1).strip()
        if presidency_member != last_presidency:
            current_header = f"District {district_counter}, {presidency_member}"
            last_presidency = presidency_member
            district_counter += 1
        continue

    if line != "":
        current_block.append(line)

if current_block:
    companionship_blocks.append((current_header, current_block))

print("Detected companionship blocks:", len(companionship_blocks))

invalid_name_keywords = {"Inc", "Corp", "LLC", "Company", "Reserve", "Tx", "Texas", "Arlington, TX"}
valid_name_pattern = re.compile(r'^[A-Z][a-z]+, [A-Z][a-zA-Z\-\.]+$')
def is_valid_name_strict(name):
    return (
        valid_name_pattern.match(name)
        and not any(k in name for k in invalid_name_keywords)
    )

final_rows = []
seen_headers = set()
style_rows = []
red_fill_rows = []

def find_do_not_contact_names(block):
    dnc_names = set()
    # Track possible simple headers to match with full names later
    header_to_fullname = {}
    fullnames = []

    # Precollect full names
    for line in block:
        match = re.match(r'^([A-Z][a-z]+),\s+([A-Z][a-zA-Z\-\.]+)$', line.strip())
        if match:
            fullnames.append(line.strip())

    # Check all lines
    for i, line in enumerate(block):
        l = line.lower()
        if "do not contact" in l:
            # Scan upward for header-like names
            for j in range(i-1, max(-1, i-5), -1):
                header_line = block[j].strip()
                if re.match(r'^[A-Z][a-z]+$', header_line):
                    for fn in fullnames:
                        if fn.startswith(header_line + ","):
                            dnc_names.add(fn)
                            break
                elif valid_name_pattern.match(header_line):
                    dnc_names.add(header_line)
                    break
    return dnc_names

def flush_final_all(header, companions, families, dnc_set):
    if header not in seen_headers:
        final_rows.append([header, "", ""])
        style_rows.append(len(final_rows))  # style this header row
        final_rows.append(["", "", ""])
        seen_headers.add(header)
    if companions:
        row_index = len(final_rows) + 1
        final_rows.append(["/".join(companions), "", ""])
        style_rows.append(row_index)  # style the companionship row
        final_rows.append(["", "", ""])
    for i, f in enumerate(families):
        comment = "Comments" if i == 0 else ""
        if f in dnc_set:
            f += "\nDo Not Contact"
            final_rows.append(["", f, comment])
            red_fill_rows.append(len(final_rows))
        else:
            final_rows.append(["", f, comment])
    final_rows.append(["", "", ""])

for header, block in companionship_blocks:
    dnc_names = find_do_not_contact_names(block)
    cleaned = [
        line.strip() for line in block
        if not re.search(r'Presid.{0,10}Member', line, re.IGNORECASE)
        and "@" not in line and not re.search(r'\d{3}-\d{3}-\d{4}', line)
    ]
    text = "\n".join(cleaned)
    names = re.findall(r'([A-Z][a-z]+, [A-Z][a-zA-Z\-\.]+)', text)
    filtered = [name for name in dict.fromkeys(names) if is_valid_name_strict(name)]

    companions = []
    families = []
    if len(filtered) >= 2:
        companions = filtered[:2]
        families = filtered[2:]
    elif len(filtered) == 1:
        companions = [filtered[0]]
        families = []
    else:
        continue
    flush_final_all(header, companions, families, dnc_names)
    dnc_names = find_do_not_contact_names(block)
    cleaned = [
        line.strip() for line in block
        if not re.search(r'Presid.{0,10}Member', line, re.IGNORECASE)
        and "@" not in line and not re.search(r'\d{3}-\d{3}-\d{4}', line)
    ]
    text = "\n".join(cleaned)
    names = re.findall(r'([A-Z][a-z]+, [A-Z][a-zA-Z\-\.]+)', text)
    filtered = []
    seen = set()
    for name in names:
        if name not in seen and is_valid_name_strict(name):
            # Only keep if not followed by other names with same last name (indent logic)
            last = name.split(",")[0]
            if sum(1 for n in names if n.startswith(last + ",")) == 1:
                filtered.append(name)
            seen.add(name)

    # Companions and families must only come from `filtered`
    companions = []
    families = []
    if len(filtered) >= 2:
        companions = filtered[:2]
        families = filtered[2:]
    elif len(filtered) == 1:
        companions = [filtered[0]]
        families = []
    else:
        continue
    flush_final_all(header, companions, families, dnc_names)

wb = Workbook()
ws = wb.active
ws.title = "Full Assignments"
for i, (a, b, c) in enumerate(final_rows, start=1):
    ws.cell(row=i, column=1, value=a)
    ws.cell(row=i, column=2, value=b)
    ws.cell(row=i, column=3, value=c)

bold_font = Font(bold=True)
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")

for row_index in style_rows:
    for col in range(1, 4):
        ws.cell(row=row_index, column=col).font = bold_font

for row_index in red_fill_rows:
    ws.cell(row=row_index, column=2).fill = red_fill

filename = generate_filename()
wb.save(filename)
print(f"Saved file: {filename}")
